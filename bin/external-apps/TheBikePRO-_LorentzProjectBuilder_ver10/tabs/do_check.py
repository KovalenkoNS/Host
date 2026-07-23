import os
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from utils.help_window import create_help_window


class DO_CheckTab:
    """Вкладка проверки DO - сравнение Excel и TXT с новой логикой"""
    
    def __init__(self, parent):
        self.parent = parent
        self.excel_file = ""
        self.txt_file = ""
        self.excel_data = []
        self.txt_data = []
        self.create_widgets()
    
    def show_help(self):
        content = """
📌 НАЗНАЧЕНИЕ

Сравнивает DO теги из Excel файла с тегами из TXT конфигурации.
Выводит ТОЛЬКО ошибки (несоответствия и отсутствующие теги).

🔍 КАК РАБОТАЕТ ПАРСИНГ TXT

1. Ищет DO блоки вида:
   _XXXX_S_SC_B01_A13_03(.ned, HV_xxx, _tag1, _tag2, ...);

2. Извлекает ТОЛЬКО настоящие теги (начинающиеся с _)
   • Теги вида .iXX - это указатели на каналы, НЕ сравниваются с Excel
   • Теги HV_ - пропускаются (это результат)
   • Строки с QUAL_STAT игнорируются

3. Каждому тегу присваивается номер канала = позиция в списке (начиная с 0)

▶️ КАК РАБОТАТЬ

1. ЗАГРУЗИТЕ EXCEL ФАЙЛ
   • Нажмите "Выбрать Excel"
   • Файл должен содержать колонки: Module, Channel, DO

2. ЗАГРУЗИТЕ TXT ФАЙЛ
   • Нажмите "Выбрать TXT"
   • Файл должен содержать DO блоки с тегами

3. Нажмите "ПРОВЕРИТЬ СООТВЕТСТВИЕ"

📊 РЕЗУЛЬТАТЫ

Показываются ТОЛЬКО ошибки:
   ❌ Теги из Excel, которых нет в TXT
   ⚠️  Несоответствия (тег есть, но канал не совпадает)
   ❌ Теги из TXT, которых нет в Excel
   ⚠️  Несоответствия TXT → Excel

📌 ОСОБЕННОСТИ

• Каналы форматируются в 2-значный формат (01, 02, ...)
• Поиск выполняется по комбинации: тег + канал
"""
        create_help_window(self.parent, "✅ Проверка DO", content)
    
    def create_widgets(self):
        # Основной фрейм с разделением на левую и правую часть
        main_paned = ttk.PanedWindow(self.parent, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True)
        
        # ============ ЛЕВАЯ ЧАСТЬ - Результаты ============
        left_frame = ttk.Frame(main_paned)
        main_paned.add(left_frame, weight=3)
        
        result_title = ttk.Label(left_frame, text="РЕЗУЛЬТАТЫ ПРОВЕРКИ DO", 
                                 font=('Arial', 12, 'bold'), foreground='blue')
        result_title.pack(pady=(0, 5))
        
        self.result_text = scrolledtext.ScrolledText(left_frame, height=20, font=("Courier", 9))
        self.result_text.pack(fill=tk.BOTH, expand=True)
        
        # Настройка стилей
        self.result_text.tag_configure("header", foreground="blue", font=('Arial', 11, 'bold'))
        self.result_text.tag_configure("error", foreground="red", font=('Courier New', 9, 'bold'))
        self.result_text.tag_configure("success", foreground="green", font=('Arial', 10, 'bold'))
        self.result_text.tag_configure("warning", foreground="purple", font=('Courier New', 9, 'bold'))
        self.result_text.tag_configure("compare", foreground="blue", font=('Courier New', 9))
        
        # ============ ПРАВАЯ ЧАСТЬ - Настройки ============
        right_frame = ttk.Frame(main_paned, width=500)
        main_paned.add(right_frame, weight=1)
        
        settings_title = ttk.Label(right_frame, text="НАСТРОЙКИ", 
                                   font=('Arial', 12, 'bold'), foreground='green')
        settings_title.pack(pady=(0, 10))
        
        # Заголовок
        header_frame = ttk.Frame(right_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        title = ttk.Label(header_frame, text="ПРОВЕРКА DO", font=("Arial", 14, "bold"))
        title.pack(side=tk.LEFT)
        
        btn_help = ttk.Button(header_frame, text="❓ Справка", command=self.show_help, width=12)
        btn_help.pack(side=tk.RIGHT)
        
        info = ttk.Label(right_frame, 
                        text="Сравнивает DO теги из Excel с TXT конфигурацией.\n👉 Нажмите 'Справка' для подробной информации.",
                        justify=tk.LEFT, font=("Arial", 9), foreground="gray", wraplength=450)
        info.pack(pady=(0, 10))
        
        # Блок загрузки файлов
        file_frame = ttk.LabelFrame(right_frame, text="Загрузка файлов", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Excel
        excel_frame = ttk.Frame(file_frame)
        excel_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(excel_frame, text="📊 Excel:", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        self.excel_path_var = tk.StringVar(value="Файл не выбран")
        ttk.Label(excel_frame, textvariable=self.excel_path_var, foreground="gray").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        btn_excel = ttk.Button(excel_frame, text="Выбрать", command=self.select_excel_file)
        btn_excel.pack(side=tk.RIGHT)
        
        # TXT
        txt_frame = ttk.Frame(file_frame)
        txt_frame.pack(fill=tk.X, pady=(5, 0))
        
        ttk.Label(txt_frame, text="📄 TXT:", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        self.txt_path_var = tk.StringVar(value="Файл не выбран")
        ttk.Label(txt_frame, textvariable=self.txt_path_var, foreground="gray").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        btn_txt = ttk.Button(txt_frame, text="Выбрать", command=self.select_txt_file)
        btn_txt.pack(side=tk.RIGHT)
        
        # Информация о данных
        info_frame = ttk.LabelFrame(right_frame, text="Информация о данных", padding="10")
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.info_excel = ttk.Label(info_frame, text="Excel: не загружен", foreground="gray")
        self.info_excel.pack(anchor=tk.W)
        
        self.info_txt = ttk.Label(info_frame, text="TXT: не загружен", foreground="gray")
        self.info_txt.pack(anchor=tk.W)
        
        # Кнопки действий
        action_frame = ttk.LabelFrame(right_frame, text="Действия", padding="10")
        action_frame.pack(fill=tk.X, pady=(0, 10))
        
        btn_check = ttk.Button(action_frame, text="🔍 ПРОВЕРИТЬ СООТВЕТСТВИЕ", command=self.check_matches)
        btn_check.pack(fill=tk.X, pady=(0, 5))
        
        btn_clear = ttk.Button(action_frame, text="ОЧИСТИТЬ", command=self.clear_all)
        btn_clear.pack(fill=tk.X, pady=(0, 5))
        
        btn_save = ttk.Button(action_frame, text="СОХРАНИТЬ ОТЧЕТ", command=self.save_report)
        btn_save.pack(fill=tk.X)
        
        # Статус
        status_frame = ttk.LabelFrame(right_frame, text="Статус", padding="10")
        status_frame.pack(fill=tk.X)
        
        self.status_label = ttk.Label(status_frame, text="Готов к работе", foreground="gray", wraplength=450)
        self.status_label.pack(anchor=tk.W)
    
    def center_window(self, window):
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f'{width}x{height}+{x}+{y}')
    
    def normalize_module(self, module):
        return module.replace('_', '').replace('-', '')
    
    def extract_tag_from_excel_line(self, line):
        match = re.search(r':=\s*([a-zA-Z0-9_]+)\s*;', line)
        if match:
            return match.group(1)
        match = re.search(r':=\s*([a-zA-Z0-9_]+)', line)
        if match:
            return match.group(1)
        return None
    
    def select_excel_file(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Ошибка", "Библиотека openpyxl не установлена!\nУстановите: pip install openpyxl")
            return
        file_path = filedialog.askopenfilename(
            title="Выберите Excel файл",
            filetypes=[
                ("Все файлы", "*.*"),
                ("Excel files", "*.xlsx *.xls")
            ]
        )
        if file_path:
            self.excel_file = file_path
            self.excel_path_var.set(os.path.basename(file_path))
            self.status_label.config(text=f"Загружен Excel: {os.path.basename(file_path)}", foreground="green")
            self.load_excel_data()
    
    def load_excel_data(self):
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
            sheet = workbook.active
            self.excel_data = []
            
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append("")
            
            module_col = None
            channel_col = None
            do_col = None
            module_keywords = ['Module', 'Модуль', 'MODULE', 'module']
            channel_keywords = ['Channel', 'Канал', 'CHANNEL', 'channel', 'Ch', 'ch']
            do_keywords = ['DO', 'DCS DO', 'Value', 'Expression', 'Signal', 'DO Expression', 'DCS Signal', 'DO Signal']
            
            for idx, header in enumerate(headers):
                header_lower = header.lower()
                if module_col is None:
                    for keyword in module_keywords:
                        if keyword.lower() in header_lower:
                            module_col = idx
                            break
                if channel_col is None:
                    for keyword in channel_keywords:
                        if keyword.lower() in header_lower:
                            channel_col = idx
                            break
                if do_col is None:
                    for keyword in do_keywords:
                        if keyword.lower() in header_lower:
                            do_col = idx
                            break
            
            if do_col is None:
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=20, values_only=True)):
                    if row:
                        for idx, cell in enumerate(row):
                            if cell and isinstance(cell, str) and ':=' in cell:
                                do_col = idx
                                break
                        if do_col is not None:
                            break
            
            if do_col is None:
                max_col = 0
                for row in sheet.iter_rows(min_row=2, max_row=10, values_only=True):
                    if row:
                        for idx, cell in enumerate(row):
                            if cell and str(cell).strip() and str(cell).strip() != 'None':
                                if idx > max_col:
                                    max_col = idx
                do_col = max_col
            
            if module_col is None:
                module_col = 0
            if channel_col is None:
                channel_col = 1 if len(headers) > 1 else 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) > max(module_col, channel_col, do_col):
                    module = str(row[module_col]).strip() if row[module_col] is not None else ""
                    channel = str(row[channel_col]).strip() if row[channel_col] is not None else ""
                    do_value = str(row[do_col]).strip() if row[do_col] is not None else ""
                    
                    if not do_value or do_value == 'None':
                        continue
                    
                    if ':=' in do_value:
                        tag = self.extract_tag_from_excel_line(do_value)
                        if tag:
                            try:
                                channel_num = int(channel)
                                channel_formatted = str(channel_num).zfill(2)
                            except (ValueError, TypeError):
                                channel_formatted = channel.zfill(2) if channel.isdigit() else channel
                            
                            self.excel_data.append({
                                'tag': tag, 'full': do_value, 'module': module,
                                'module_normalized': self.normalize_module(module),
                                'channel': channel, 'channel_formatted': channel_formatted,
                                'channel_num': int(channel) if channel.isdigit() else None
                            })
            
            self.info_excel.config(text=f"Excel: загружено {len(self.excel_data)} записей", foreground="green")
            self.status_label.config(text=f"Загружено {len(self.excel_data)} записей из Excel", foreground="green")
            if len(self.excel_data) == 0:
                messagebox.showwarning("Предупреждение", "Не найдено данных в Excel файле.")
            workbook.close()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить Excel файл:\n{str(e)}")
            self.excel_data = []
    
    def select_txt_file(self):
        file_path = filedialog.askopenfilename(
            title="Выберите TXT файл",
            filetypes=[
                ("Все файлы", "*.*"),
                ("Text files", "*.txt")
            ]
        )
        if file_path:
            self.txt_file = file_path
            self.txt_path_var.set(os.path.basename(file_path))
            self.status_label.config(text=f"Загружен TXT: {os.path.basename(file_path)}", foreground="green")
            self.load_txt_data()
    
    def load_txt_data(self):
        try:
            self.txt_data = []
            with open(self.txt_file, 'r', encoding='utf-8') as file:
                content = file.read()
            
            matches = []
            
            pattern1 = r'(_[A-Za-z0-9_]+)\s*\(\s*\.ned\s*,\s*([^;]+?)\s*\)\s*;'
            matches1 = re.findall(pattern1, content, re.DOTALL | re.IGNORECASE)
            if matches1:
                matches.extend(matches1)
            
            if not matches:
                pattern2 = r'(_[A-Za-z0-9_]+)\s*\(\s*([^;]+?)\s*\)\s*;'
                matches2 = re.findall(pattern2, content, re.DOTALL | re.IGNORECASE)
                for func_name, args_str in matches2:
                    if 'HV_' in args_str:
                        matches.append((func_name, args_str))
            
            if not matches:
                pattern3 = r'(_[A-Za-z0-9_]+)\s*\(\s*\.ned\s*,\s*([\s\S]+?)\s*\)\s*;'
                matches3 = re.findall(pattern3, content, re.DOTALL | re.IGNORECASE)
                if matches3:
                    matches.extend(matches3)
            
            if not matches:
                messagebox.showwarning("Предупреждение", "Не найдено DO блоков в TXT файле.")
                return
            
            processed_functions = []
            for func_name, args_str in matches:
                args = []
                current_arg = ""
                in_comment = False
                in_parentheses = 0
                
                for char in args_str:
                    if char == '(' and not in_comment:
                        in_parentheses += 1
                        current_arg += char
                    elif char == ')' and not in_comment:
                        in_parentheses -= 1
                        current_arg += char
                    elif char == '(' and in_comment:
                        current_arg += char
                    elif char == ')' and in_comment:
                        current_arg += char
                    elif char == ',' and not in_comment and in_parentheses == 0:
                        args.append(current_arg.strip())
                        current_arg = ""
                    else:
                        current_arg += char
                if current_arg.strip():
                    args.append(current_arg.strip())
                
                args = [arg.strip() for arg in args if arg.strip()]
                
                hv_index = -1
                for i, arg in enumerate(args):
                    if arg.startswith('HV_'):
                        hv_index = i
                        break
                if hv_index >= 0:
                    args.pop(hv_index)
                
                if not args:
                    continue
                
                for channel_num, arg in enumerate(args, start=0):
                    if not arg:
                        continue
                    if arg.startswith('.i'):
                        continue
                    
                    tag_name = arg.split('(*')[0].strip()
                    if not tag_name.startswith('_'):
                        continue
                    
                    self.txt_data.append({
                        'tag': tag_name, 'channel': str(channel_num).zfill(2),
                        'channel_num': channel_num, 'position': channel_num,
                        'full_arg': arg, 'function': func_name
                    })
                processed_functions.append(func_name)
            
            self.info_txt.config(text=f"TXT: загружено {len(self.txt_data)} тегов из {len(processed_functions)} DO блоков", foreground="green")
            self.status_label.config(text=f"Загружено {len(self.txt_data)} тегов из TXT", foreground="green")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить TXT файл:\n{str(e)}")
            self.txt_data = []
    
    def check_matches(self):
        if not self.excel_data:
            messagebox.showwarning("Предупреждение", "Сначала загрузите Excel файл!")
            return
        if not self.txt_data:
            messagebox.showwarning("Предупреждение", "Сначала загрузите TXT файл!")
            return
        
        self.result_text.delete('1.0', tk.END)
        
        txt_dict = {}
        for txt_item in self.txt_data:
            key = (txt_item['tag'], txt_item['channel'])
            txt_dict[key] = txt_item
        
        excel_dict = {}
        for excel_item in self.excel_data:
            key = (excel_item['tag'], excel_item['channel_formatted'])
            excel_dict[key] = excel_item
        
        excel_tags = set(item['tag'] for item in self.excel_data)
        txt_tags = set(item['tag'] for item in self.txt_data)
        
        matched_count = 0
        excel_not_in_txt = []
        excel_mismatch = []
        txt_not_in_excel = []
        txt_mismatch = []
        
        for excel_item in self.excel_data:
            tag = excel_item['tag']
            module = excel_item['module']
            channel_formatted = excel_item['channel_formatted']
            key = (tag, channel_formatted)
            
            if key in txt_dict:
                matched_count += 1
            else:
                found = False
                for txt_item in self.txt_data:
                    if txt_item['tag'] == tag:
                        found = True
                        excel_mismatch.append({
                            'tag': tag, 'excel_module': module, 'excel_channel': channel_formatted,
                            'txt_channel': txt_item['channel'], 'txt_full_arg': txt_item['full_arg'],
                            'txt_function': txt_item.get('function', 'unknown')
                        })
                        break
                if not found:
                    excel_not_in_txt.append({'tag': tag, 'excel_module': module, 'excel_channel': channel_formatted})
        
        for txt_item in self.txt_data:
            tag = txt_item['tag']
            channel = txt_item['channel']
            key = (tag, channel)
            
            if key in excel_dict:
                pass
            else:
                found = False
                for excel_item in self.excel_data:
                    if excel_item['tag'] == tag:
                        found = True
                        txt_mismatch.append({
                            'tag': tag, 'txt_channel': channel, 'txt_full_arg': txt_item['full_arg'],
                            'txt_function': txt_item.get('function', 'unknown'),
                            'excel_channel': excel_item['channel_formatted'], 'excel_module': excel_item['module']
                        })
                        break
                if not found:
                    txt_not_in_excel.append({
                        'tag': tag, 'txt_channel': channel, 'txt_full_arg': txt_item['full_arg'],
                        'txt_function': txt_item.get('function', 'unknown')
                    })
        
        total_errors = len(excel_not_in_txt) + len(excel_mismatch) + len(txt_not_in_excel) + len(txt_mismatch)
        
        if total_errors == 0:
            self.result_text.insert(tk.END, "="*80 + "\n")
            self.result_text.insert(tk.END, "🎉 ВСЕ ТЕГИ СОВПАДАЮТ! ОШИБОК НЕ НАЙДЕНО.\n", "success")
            self.result_text.insert(tk.END, "="*80 + "\n")
            self.status_label.config(text="✅ Все теги совпадают! Ошибок нет.", foreground="green")
            messagebox.showinfo("Успех", "Все теги совпадают! Ошибок не найдено.")
            return
        
        self.result_text.insert(tk.END, "="*80 + "\n")
        self.result_text.insert(tk.END, "РЕЗУЛЬТАТЫ ПРОВЕРКИ DO\n", "header")
        self.result_text.insert(tk.END, "="*80 + "\n\n")
        self.result_text.insert(tk.END, f"⚠️  НАЙДЕНО {total_errors} ОШИБОК(И)\n", "error")
        self.result_text.insert(tk.END, "="*80 + "\n\n")
        
        self.result_text.insert(tk.END, "📊 СТАТИСТИКА ПО ТЕГАМ:\n", "header")
        self.result_text.insert(tk.END, f"   Всего тегов в Excel: {len(excel_tags)}\n")
        self.result_text.insert(tk.END, f"   Всего тегов в TXT:   {len(txt_tags)}\n")
        self.result_text.insert(tk.END, f"   Общих тегов:         {len(excel_tags & txt_tags)}\n")
        self.result_text.insert(tk.END, f"   Только в Excel:      {len(excel_tags - txt_tags)}\n")
        self.result_text.insert(tk.END, f"   Только в TXT:        {len(txt_tags - excel_tags)}\n\n")
        
        if excel_not_in_txt:
            self.result_text.insert(tk.END, "="*80 + "\n")
            self.result_text.insert(tk.END, f"❌ ТЕГИ ИЗ EXCEL, КОТОРЫХ НЕТ В TXT ({len(excel_not_in_txt)}):\n", "error")
            self.result_text.insert(tk.END, "="*80 + "\n\n")
            for item in excel_not_in_txt:
                self.result_text.insert(tk.END, f"  ❌ {item['tag']}\n", "error")
                self.result_text.insert(tk.END, f"     Excel: модуль={item['excel_module']}, канал={item['excel_channel']}\n", "compare")
                self.result_text.insert(tk.END, f"     TXT:  НЕ НАЙДЕН\n", "compare")
                self.result_text.insert(tk.END, "-"*80 + "\n")
            self.result_text.insert(tk.END, "\n")
        
        if excel_mismatch:
            self.result_text.insert(tk.END, "="*80 + "\n")
            self.result_text.insert(tk.END, f"⚠️  НЕСООТВЕТСТВИЯ Excel → TXT (тег есть, канал не совпадает) ({len(excel_mismatch)}):\n", "warning")
            self.result_text.insert(tk.END, "="*80 + "\n\n")
            for item in excel_mismatch:
                self.result_text.insert(tk.END, f"  ⚠️  {item['tag']}\n", "warning")
                self.result_text.insert(tk.END, f"     Excel: модуль={item['excel_module']}, канал={item['excel_channel']}\n", "compare")
                self.result_text.insert(tk.END, f"     TXT:  канал={item['txt_channel']}, аргумент={item['txt_full_arg']}\n", "compare")
                if item.get('txt_function'):
                    self.result_text.insert(tk.END, f"     Функция: {item['txt_function']}\n", "compare")
                self.result_text.insert(tk.END, "-"*80 + "\n")
            self.result_text.insert(tk.END, "\n")
        
        if txt_not_in_excel:
            self.result_text.insert(tk.END, "="*80 + "\n")
            self.result_text.insert(tk.END, f"❌ ТЕГИ ИЗ TXT, КОТОРЫХ НЕТ В EXCEL ({len(txt_not_in_excel)}):\n", "error")
            self.result_text.insert(tk.END, "="*80 + "\n\n")
            for item in txt_not_in_excel:
                self.result_text.insert(tk.END, f"  ❌ {item['tag']}\n", "error")
                self.result_text.insert(tk.END, f"     TXT:  канал={item['txt_channel']}, аргумент={item['txt_full_arg']}\n", "compare")
                if item.get('txt_function'):
                    self.result_text.insert(tk.END, f"     Функция: {item['txt_function']}\n", "compare")
                self.result_text.insert(tk.END, f"     Excel: НЕ НАЙДЕН\n", "compare")
                self.result_text.insert(tk.END, "-"*80 + "\n")
            self.result_text.insert(tk.END, "\n")
        
        if txt_mismatch:
            self.result_text.insert(tk.END, "="*80 + "\n")
            self.result_text.insert(tk.END, f"⚠️  НЕСООТВЕТСТВИЯ TXT → Excel (тег есть, канал не совпадает) ({len(txt_mismatch)}):\n", "warning")
            self.result_text.insert(tk.END, "="*80 + "\n\n")
            for item in txt_mismatch:
                self.result_text.insert(tk.END, f"  ⚠️  {item['tag']}\n", "warning")
                self.result_text.insert(tk.END, f"     TXT:  канал={item['txt_channel']}, аргумент={item['txt_full_arg']}\n", "compare")
                if item.get('txt_function'):
                    self.result_text.insert(tk.END, f"     Функция: {item['txt_function']}\n", "compare")
                self.result_text.insert(tk.END, f"     Excel: модуль={item.get('excel_module', 'unknown')}, канал={item['excel_channel']}\n", "compare")
                self.result_text.insert(tk.END, "-"*80 + "\n")
            self.result_text.insert(tk.END, "\n")
        
        self.result_text.insert(tk.END, "="*80 + "\n")
        self.result_text.insert(tk.END, "ИТОГОВАЯ СТАТИСТИКА:\n", "header")
        self.result_text.insert(tk.END, "="*80 + "\n\n")
        self.result_text.insert(tk.END, f"📊 Всего записей:\n")
        self.result_text.insert(tk.END, f"   Excel: {len(self.excel_data)} записей\n")
        self.result_text.insert(tk.END, f"   TXT:   {len(self.txt_data)} тегов\n\n")
        self.result_text.insert(tk.END, f"📊 Найдено ошибок:\n")
        self.result_text.insert(tk.END, f"   ❌ В Excel, но не в TXT:  {len(excel_not_in_txt)}\n")
        self.result_text.insert(tk.END, f"   ⚠️  Несоответствия Excel→TXT: {len(excel_mismatch)}\n")
        self.result_text.insert(tk.END, f"   ❌ В TXT, но не в Excel:  {len(txt_not_in_excel)}\n")
        self.result_text.insert(tk.END, f"   ⚠️  Несоответствия TXT→Excel: {len(txt_mismatch)}\n")
        self.result_text.insert(tk.END, f"\n   ✅ Совпадают (тег+канал): {matched_count}\n", "success")
        self.result_text.insert(tk.END, f"   ⚠️  Всего ошибок: {total_errors}\n", "error")
        self.result_text.insert(tk.END, "\n" + "="*80 + "\n")
        
        self.status_label.config(text=f"⚠️ Найдено {total_errors} ошибок", foreground="red")
        messagebox.showwarning("Результаты проверки",
            f"⚠️  НАЙДЕНО {total_errors} ОШИБОК(И)\n\n"
            f"📊 Статистика по тегам:\n   Общих тегов: {len(excel_tags & txt_tags)}\n"
            f"   Только в Excel: {len(excel_tags - txt_tags)}\n   Только в TXT: {len(txt_tags - excel_tags)}\n\n"
            f"   ❌ В Excel, но не в TXT: {len(excel_not_in_txt)}\n"
            f"   ⚠️  Несоответствия Excel→TXT: {len(excel_mismatch)}\n"
            f"   ❌ В TXT, но не в Excel: {len(txt_not_in_excel)}\n"
            f"   ⚠️  Несоответствия TXT→Excel: {len(txt_mismatch)}\n\n"
            f"✅ Совпадают (тег+канал): {matched_count}")
    
    def clear_all(self):
        self.excel_file = ""
        self.txt_file = ""
        self.excel_data = []
        self.txt_data = []
        self.excel_path_var.set("Файл не выбран")
        self.txt_path_var.set("Файл не выбран")
        self.result_text.delete('1.0', tk.END)
        self.info_excel.config(text="Excel: не загружен", foreground="gray")
        self.info_txt.config(text="TXT: не загружен", foreground="gray")
        self.status_label.config(text="Очищено", foreground="gray")
    
    def save_report(self):
        content = self.result_text.get('1.0', tk.END).strip()
        if not content:
            messagebox.showwarning("Предупреждение", "Нет данных для сохранения!")
            return
        file_path = filedialog.asksaveasfilename(
            title="Сохранить отчет",
            defaultextension=".txt",
            filetypes=[
                ("Текстовые файлы", "*.txt"),
                ("Все файлы", "*.*")
            ]
        )
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                self.status_label.config(text=f"Отчет сохранен: {os.path.basename(file_path)}", foreground="green")
                messagebox.showinfo("Успех", f"Отчет сохранен в:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить отчет:\n{str(e)}")