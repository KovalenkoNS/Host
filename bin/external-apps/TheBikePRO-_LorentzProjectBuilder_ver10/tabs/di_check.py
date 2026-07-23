import os
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from utils.help_window import create_help_window


class DI_DOCheckTab:
    """Вкладка проверки DI - сравнение Excel и TXT файлов"""
    
    def __init__(self, parent):
        self.parent = parent
        self.excel_file = ""
        self.txt_file = ""
        self.excel_data = []
        self.txt_data = []
        self.mismatches = []
        self.create_widgets()
    
    def show_help(self):
        content = """
📌 НАЗНАЧЕНИЕ

Сравнивает теги из Excel файла (DI) с тегами из TXT конфигурации.
Выводит ТОЛЬКО ошибки (несоответствия и отсутствующие теги).

▶️ КАК РАБОТАТЬ

1. ЗАГРУЗИТЕ EXCEL ФАЙЛ
   • Нажмите "Выбрать Excel"
   • Файл должен содержать колонки: DCS DI, Module, Channel

2. ЗАГРУЗИТЕ TXT ФАЙЛ
   • Нажмите "Выбрать TXT"
   • Файл должен содержать строки с тегами вида: _tag := value;

3. Нажмите "ПРОВЕРИТЬ СООТВЕТСТВИЕ"

📊 РЕЗУЛЬТАТЫ

Показываются ТОЛЬКО ошибки:
   ❌ Теги из Excel, которых нет в TXT
   ⚠️  Несоответствия (тег есть, но модуль/канал не совпадают)
   ❌ Теги из TXT, которых нет в Excel
   ⚠️  Несоответствия TXT → Excel

📌 ОСОБЕННОСТИ

• Имена модулей нормализуются (удаляются _ и -)
• Каналы форматируются в 2-значный формат (01, 02, ...)
• Поиск выполняется по комбинации: тег + нормализованный модуль + канал
"""
        create_help_window(self.parent, "✅ Проверка DI", content)
    
    def create_widgets(self):
        # Основной фрейм с разделением на левую и правую часть
        main_paned = ttk.PanedWindow(self.parent, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True)
        
        # ============ ЛЕВАЯ ЧАСТЬ - Результаты ============
        left_frame = ttk.Frame(main_paned)
        main_paned.add(left_frame, weight=3)
        
        result_title = ttk.Label(left_frame, text="РЕЗУЛЬТАТЫ ПРОВЕРКИ DI", 
                                 font=('Arial', 12, 'bold'), foreground='blue')
        result_title.pack(pady=(0, 5))
        
        self.result_text = scrolledtext.ScrolledText(left_frame, height=20, font=("Courier", 9))
        self.result_text.pack(fill=tk.BOTH, expand=True)
        
        # Настройка стилей
        self.result_text.tag_configure("header", foreground="blue", font=('Arial', 11, 'bold'))
        self.result_text.tag_configure("error", foreground="red", font=('Courier New', 9, 'bold'))
        self.result_text.tag_configure("success", foreground="green", font=('Arial', 10, 'bold'))
        self.result_text.tag_configure("warning", foreground="purple", font=('Courier New', 9, 'bold'))
        self.result_text.tag_configure("compare", foreground="blue", font=('Courier New', 9))
        
        # ============ ПРАВАЯ ЧАСТЬ - Настройки ============
        right_frame = ttk.Frame(main_paned, width=500)
        main_paned.add(right_frame, weight=1)
        
        settings_title = ttk.Label(right_frame, text="НАСТРОЙКИ", 
                                   font=('Arial', 12, 'bold'), foreground='green')
        settings_title.pack(pady=(0, 10))
        
        # Заголовок
        header_frame = ttk.Frame(right_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        title = ttk.Label(header_frame, text="ПРОВЕРКА DI", font=("Arial", 14, "bold"))
        title.pack(side=tk.LEFT)
        
        btn_help = ttk.Button(header_frame, text="❓ Справка", command=self.show_help, width=12)
        btn_help.pack(side=tk.RIGHT)
        
        info = ttk.Label(right_frame, 
                        text="Сравнивает теги из Excel (DI) с TXT конфигурацией.\n👉 Нажмите 'Справка' для подробной информации.",
                        justify=tk.LEFT, font=("Arial", 9), foreground="gray", wraplength=450)
        info.pack(pady=(0, 10))
        
        # Блок загрузки файлов
        file_frame = ttk.LabelFrame(right_frame, text="Загрузка файлов", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Excel
        excel_frame = ttk.Frame(file_frame)
        excel_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(excel_frame, text="📊 Excel:", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        self.excel_path_var = tk.StringVar(value="Файл не выбран")
        ttk.Label(excel_frame, textvariable=self.excel_path_var, foreground="gray").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        btn_excel = ttk.Button(excel_frame, text="Выбрать", command=self.select_excel_file)
        btn_excel.pack(side=tk.RIGHT)
        
        # TXT
        txt_frame = ttk.Frame(file_frame)
        txt_frame.pack(fill=tk.X, pady=(5, 0))
        
        ttk.Label(txt_frame, text="📄 TXT:", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        self.txt_path_var = tk.StringVar(value="Файл не выбран")
        ttk.Label(txt_frame, textvariable=self.txt_path_var, foreground="gray").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        btn_txt = ttk.Button(txt_frame, text="Выбрать", command=self.select_txt_file)
        btn_txt.pack(side=tk.RIGHT)
        
        # Информация о данных
        info_frame = ttk.LabelFrame(right_frame, text="Информация о данных", padding="10")
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.info_excel = ttk.Label(info_frame, text="Excel: не загружен", foreground="gray")
        self.info_excel.pack(anchor=tk.W)
        
        self.info_txt = ttk.Label(info_frame, text="TXT: не загружен", foreground="gray")
        self.info_txt.pack(anchor=tk.W)
        
        # Кнопки действий
        action_frame = ttk.LabelFrame(right_frame, text="Действия", padding="10")
        action_frame.pack(fill=tk.X, pady=(0, 10))
        
        btn_check = ttk.Button(action_frame, text="🔍 ПРОВЕРИТЬ СООТВЕТСТВИЕ", command=self.check_matches)
        btn_check.pack(fill=tk.X, pady=(0, 5))
        
        btn_clear = ttk.Button(action_frame, text="ОЧИСТИТЬ", command=self.clear_all)
        btn_clear.pack(fill=tk.X, pady=(0, 5))
        
        btn_save = ttk.Button(action_frame, text="СОХРАНИТЬ ОТЧЕТ", command=self.save_report)
        btn_save.pack(fill=tk.X)
        
        # Статус
        status_frame = ttk.LabelFrame(right_frame, text="Статус", padding="10")
        status_frame.pack(fill=tk.X)
        
        self.status_label = ttk.Label(status_frame, text="Готов к работе", foreground="gray", wraplength=450)
        self.status_label.pack(anchor=tk.W)
    
    def center_window(self, window):
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f'{width}x{height}+{x}+{y}')
    
    def normalize_module(self, module):
        return module.replace('_', '').replace('-', '')
    
    def extract_tag_from_line(self, line):
        match = re.search(r'(_[a-zA-Z0-9_\.]+)\s*:=', line)
        if match:
            return match.group(1)
        match = re.search(r'(_[a-zA-Z0-9_\.]+)', line)
        if match:
            tag = match.group(1)
            end_pos = match.end()
            if end_pos < len(line):
                next_chars = line[end_pos:end_pos+3].strip()
                if next_chars.startswith(':=') or next_chars == '':
                    return tag
        return None
    
    def select_excel_file(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Ошибка", "Библиотека openpyxl не установлена!\nУстановите: pip install openpyxl")
            return
        file_path = filedialog.askopenfilename(
            title="Выберите Excel файл",
            filetypes=[
                ("Все файлы", "*.*"),
                ("Excel files", "*.xlsx *.xls")
            ]
        )
        if file_path:
            self.excel_file = file_path
            self.excel_path_var.set(os.path.basename(file_path))
            self.status_label.config(text=f"Загружен Excel: {os.path.basename(file_path)}", foreground="green")
            self.load_excel_data()
    
    def load_excel_data(self):
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
            sheet = workbook.active
            self.excel_data = []
            dcs_di_col = None
            module_col = None
            channel_col = None
            
            for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
                for idx, cell in enumerate(row):
                    if cell:
                        cell_str = str(cell).strip()
                        if 'DCS DI' in cell_str or 'DI' in cell_str:
                            dcs_di_col = idx
                        elif 'Module' in cell_str or 'Модуль' in cell_str:
                            module_col = idx
                        elif 'Channel' in cell_str or 'Канал' in cell_str:
                            channel_col = idx
                if dcs_di_col is not None:
                    break
            
            if dcs_di_col is None:
                dcs_di_col = 0
                module_col = 1
                channel_col = 2
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) > max(dcs_di_col or 0, module_col or 0, channel_col or 0):
                    dcs_di = str(row[dcs_di_col]).strip() if row[dcs_di_col] is not None else ""
                    module = str(row[module_col]).strip() if row[module_col] is not None else ""
                    channel = str(row[channel_col]).strip() if row[channel_col] is not None else ""
                    if dcs_di and dcs_di != 'None':
                        if ':' in dcs_di:
                            tag = dcs_di.split(':')[0].strip()
                        else:
                            tag = dcs_di
                        if '=' in tag:
                            tag = tag.split('=')[0].strip()
                        channel_formatted = channel.zfill(2) if channel.isdigit() else channel
                        self.excel_data.append({
                            'tag': tag, 'full': dcs_di, 'module': module,
                            'module_normalized': self.normalize_module(module),
                            'channel': channel, 'channel_formatted': channel_formatted
                        })
            
            self.info_excel.config(text=f"Excel: загружено {len(self.excel_data)} записей", foreground="green")
            self.status_label.config(text=f"Загружено {len(self.excel_data)} записей из Excel", foreground="green")
            workbook.close()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить Excel файл:\n{str(e)}")
            self.excel_data = []
    
    def select_txt_file(self):
        file_path = filedialog.askopenfilename(
            title="Выберите TXT файл",
            filetypes=[
                ("Все файлы", "*.*"),
                ("Text files", "*.txt")
            ]
        )
        if file_path:
            self.txt_file = file_path
            self.txt_path_var.set(os.path.basename(file_path))
            self.status_label.config(text=f"Загружен TXT: {os.path.basename(file_path)}", foreground="green")
            self.load_txt_data()
    
    def load_txt_data(self):
        try:
            self.txt_data = []
            with open(self.txt_file, 'r', encoding='utf-8') as file:
                lines = file.readlines()
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                tag = self.extract_tag_from_line(line)
                if not tag:
                    continue
                match = re.search(r'([A-Za-z0-9_]+)\._([0-9]{2})', line)
                if not match:
                    match = re.search(r'_([0-9]{2})', line)
                    if match:
                        channel = match.group(1)
                        module_match = re.search(r'([A-Za-z0-9_]+)_' + channel, line)
                        if module_match:
                            full_module = module_match.group(1)
                        else:
                            full_module = "unknown"
                    else:
                        continue
                else:
                    full_module = match.group(1)
                    channel = match.group(2)
                
                module_parts = full_module.split('_')
                if len(module_parts) >= 2:
                    module = '_'.join(module_parts[-2:]) if len(module_parts) >= 2 else module_parts[-1]
                else:
                    module = full_module
                
                self.txt_data.append({
                    'tag': tag, 'module': module,
                    'module_normalized': self.normalize_module(module),
                    'channel': channel, 'full_line': line, 'full_module': full_module
                })
            
            self.info_txt.config(text=f"TXT: загружено {len(self.txt_data)} записей", foreground="green")
            self.status_label.config(text=f"Загружено {len(self.txt_data)} записей из TXT", foreground="green")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить TXT файл:\n{str(e)}")
            self.txt_data = []
    
    def check_matches(self):
        if not self.excel_data:
            messagebox.showwarning("Предупреждение", "Сначала загрузите Excel файл!")
            return
        if not self.txt_data:
            messagebox.showwarning("Предупреждение", "Сначала загрузите TXT файл!")
            return
        
        self.result_text.delete('1.0', tk.END)
        
        txt_dict = {}
        for txt_item in self.txt_data:
            key = (txt_item['tag'], txt_item['module_normalized'], txt_item['channel'])
            txt_dict[key] = txt_item
        
        excel_dict = {}
        for excel_item in self.excel_data:
            key = (excel_item['tag'], excel_item['module_normalized'], excel_item['channel_formatted'])
            excel_dict[key] = excel_item
        
        matched_count = 0
        excel_not_in_txt = []
        excel_mismatch = []
        txt_not_in_excel = []
        txt_mismatch = []
        
        for excel_item in self.excel_data:
            tag = excel_item['tag']
            module = excel_item['module']
            module_normalized = excel_item['module_normalized']
            channel = excel_item['channel']
            channel_formatted = excel_item['channel_formatted']
            key = (tag, module_normalized, channel_formatted)
            
            if key in txt_dict:
                matched_count += 1
            else:
                found = False
                for txt_item in self.txt_data:
                    if txt_item['tag'] == tag:
                        found = True
                        excel_mismatch.append({
                            'tag': tag, 'excel_module': module, 'excel_channel': channel,
                            'txt_module': txt_item['module'], 'txt_channel': txt_item['channel']
                        })
                        break
                if not found:
                    excel_not_in_txt.append({'tag': tag, 'excel_module': module, 'excel_channel': channel})
        
        for txt_item in self.txt_data:
            tag = txt_item['tag']
            module = txt_item['module']
            module_normalized = txt_item['module_normalized']
            channel = txt_item['channel']
            key = (tag, module_normalized, channel)
            
            if key in excel_dict:
                pass
            else:
                found = False
                for excel_item in self.excel_data:
                    if excel_item['tag'] == tag:
                        found = True
                        txt_mismatch.append({
                            'tag': tag, 'txt_module': module, 'txt_channel': channel,
                            'excel_module': excel_item['module'], 'excel_channel': excel_item['channel']
                        })
                        break
                if not found:
                    txt_not_in_excel.append({'tag': tag, 'txt_module': module, 'txt_channel': channel})
        
        total_errors = len(excel_not_in_txt) + len(excel_mismatch) + len(txt_not_in_excel) + len(txt_mismatch)
        
        if total_errors == 0:
            self.result_text.insert(tk.END, "="*80 + "\n")
            self.result_text.insert(tk.END, "🎉 ВСЕ ТЕГИ СОВПАДАЮТ! ОШИБОК НЕ НАЙДЕНО.\n", "success")
            self.result_text.insert(tk.END, "="*80 + "\n")
            self.status_label.config(text="✅ Все теги совпадают! Ошибок нет.", foreground="green")
            messagebox.showinfo("Успех", "Все теги совпадают! Ошибок не найдено.")
            return
        
        self.result_text.insert(tk.END, "="*80 + "\n")
        self.result_text.insert(tk.END, "РЕЗУЛЬТАТЫ ПРОВЕРКИ DI\n", "header")
        self.result_text.insert(tk.END, "="*80 + "\n\n")
        self.result_text.insert(tk.END, f"⚠️  НАЙДЕНО {total_errors} ОШИБОК(И)\n", "error")
        self.result_text.insert(tk.END, "="*80 + "\n\n")
        
        if excel_not_in_txt:
            self.result_text.insert(tk.END, "="*80 + "\n")
            self.result_text.insert(tk.END, f"❌ ТЕГИ ИЗ EXCEL, КОТОРЫХ НЕТ В TXT ({len(excel_not_in_txt)}):\n", "error")
            self.result_text.insert(tk.END, "="*80 + "\n\n")
            for item in excel_not_in_txt:
                self.result_text.insert(tk.END, f"  ❌ {item['tag']}\n", "error")
                self.result_text.insert(tk.END, f"     Excel: модуль={item['excel_module']}, канал={item['excel_channel']}\n", "compare")
                self.result_text.insert(tk.END, f"     TXT:  НЕ НАЙДЕН\n", "compare")
                self.result_text.insert(tk.END, "-"*80 + "\n")
            self.result_text.insert(tk.END, "\n")
        
        if excel_mismatch:
            self.result_text.insert(tk.END, "="*80 + "\n")
            self.result_text.insert(tk.END, f"⚠️  НЕСООТВЕТСТВИЯ EXCEL → TXT ({len(excel_mismatch)}):\n", "warning")
            self.result_text.insert(tk.END, "="*80 + "\n\n")
            for item in excel_mismatch:
                self.result_text.insert(tk.END, f"  ⚠️  {item['tag']}\n", "warning")
                self.result_text.insert(tk.END, f"     Excel: модуль={item['excel_module']}, канал={item['excel_channel']}\n", "compare")
                self.result_text.insert(tk.END, f"     TXT:  модуль={item['txt_module']}, канал={item['txt_channel']}\n", "compare")
                self.result_text.insert(tk.END, "-"*80 + "\n")
            self.result_text.insert(tk.END, "\n")
        
        if txt_not_in_excel:
            self.result_text.insert(tk.END, "="*80 + "\n")
            self.result_text.insert(tk.END, f"❌ ТЕГИ ИЗ TXT, КОТОРЫХ НЕТ В EXCEL ({len(txt_not_in_excel)}):\n", "error")
            self.result_text.insert(tk.END, "="*80 + "\n\n")
            for item in txt_not_in_excel:
                self.result_text.insert(tk.END, f"  ❌ {item['tag']}\n", "error")
                self.result_text.insert(tk.END, f"     TXT:  модуль={item['txt_module']}, канал={item['txt_channel']}\n", "compare")
                self.result_text.insert(tk.END, f"     Excel: НЕ НАЙДЕН\n", "compare")
                self.result_text.insert(tk.END, "-"*80 + "\n")
            self.result_text.insert(tk.END, "\n")
        
        if txt_mismatch:
            self.result_text.insert(tk.END, "="*80 + "\n")
            self.result_text.insert(tk.END, f"⚠️  НЕСООТВЕТСТВИЯ TXT → EXCEL ({len(txt_mismatch)}):\n", "warning")
            self.result_text.insert(tk.END, "="*80 + "\n\n")
            for item in txt_mismatch:
                self.result_text.insert(tk.END, f"  ⚠️  {item['tag']}\n", "warning")
                self.result_text.insert(tk.END, f"     TXT:  модуль={item['txt_module']}, канал={item['txt_channel']}\n", "compare")
                self.result_text.insert(tk.END, f"     Excel: модуль={item['excel_module']}, канал={item['excel_channel']}\n", "compare")
                self.result_text.insert(tk.END, "-"*80 + "\n")
            self.result_text.insert(tk.END, "\n")
        
        self.result_text.insert(tk.END, "="*80 + "\n")
        self.result_text.insert(tk.END, "ИТОГОВАЯ СТАТИСТИКА:\n", "header")
        self.result_text.insert(tk.END, "="*80 + "\n\n")
        self.result_text.insert(tk.END, f"📊 Всего проверено:\n")
        self.result_text.insert(tk.END, f"   Excel: {len(self.excel_data)} записей\n")
        self.result_text.insert(tk.END, f"   TXT:   {len(self.txt_data)} записей\n\n")
        self.result_text.insert(tk.END, f"📊 Найдено ошибок:\n")
        self.result_text.insert(tk.END, f"   ❌ В Excel, но не в TXT:  {len(excel_not_in_txt)}\n")
        self.result_text.insert(tk.END, f"   ⚠️  Несоответствия Excel→TXT: {len(excel_mismatch)}\n")
        self.result_text.insert(tk.END, f"   ❌ В TXT, но не в Excel:  {len(txt_not_in_excel)}\n")
        self.result_text.insert(tk.END, f"   ⚠️  Несоответствия TXT→Excel: {len(txt_mismatch)}\n")
        self.result_text.insert(tk.END, f"\n   ✅ Совпадают: {matched_count}\n", "success")
        self.result_text.insert(tk.END, f"   ⚠️  Всего ошибок: {total_errors}\n", "error")
        self.result_text.insert(tk.END, "\n" + "="*80 + "\n")
        
        self.status_label.config(text=f"⚠️ Найдено {total_errors} ошибок", foreground="red")
        messagebox.showwarning("Результаты проверки",
            f"⚠️  НАЙДЕНО {total_errors} ОШИБОК(И)\n\n"
            f"📊 Excel:\n  ❌ Не найдено в TXT: {len(excel_not_in_txt)}\n  ⚠️  Несоответствий: {len(excel_mismatch)}\n\n"
            f"📊 TXT:\n  ❌ Не найдено в Excel: {len(txt_not_in_excel)}\n  ⚠️  Несоответствий: {len(txt_mismatch)}\n\n"
            f"✅ Совпадают: {matched_count}")
    
    def clear_all(self):
        self.excel_file = ""
        self.txt_file = ""
        self.excel_data = []
        self.txt_data = []
        self.mismatches = []
        self.excel_path_var.set("Файл не выбран")
        self.txt_path_var.set("Файл не выбран")
        self.result_text.delete('1.0', tk.END)
        self.info_excel.config(text="Excel: не загружен", foreground="gray")
        self.info_txt.config(text="TXT: не загружен", foreground="gray")
        self.status_label.config(text="Очищено", foreground="gray")
    
    def save_report(self):
        content = self.result_text.get('1.0', tk.END).strip()
        if not content:
            messagebox.showwarning("Предупреждение", "Нет данных для сохранения!")
            return
        file_path = filedialog.asksaveasfilename(
            title="Сохранить отчет",
            defaultextension=".txt",
            filetypes=[
                ("Текстовые файлы", "*.txt"),
                ("Все файлы", "*.*")
            ]
        )
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                self.status_label.config(text=f"Отчет сохранен: {os.path.basename(file_path)}", foreground="green")
                messagebox.showinfo("Успех", f"Отчет сохранен в:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить отчет:\n{str(e)}")